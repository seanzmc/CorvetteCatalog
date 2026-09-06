"""Import the pinned workbook into a disposable, independently readable SQLite DB.

No reference-project imports, generators, workbook writes, or runtime JSON inputs.
The source cells are evidence; consumers use the typed tables in schema.py.
"""
from __future__ import annotations

import argparse
from collections import Counter
from decimal import Decimal, InvalidOperation
import hashlib
import io
import json
import os
from pathlib import Path
import re
import tarfile
import tempfile
import uuid

from openpyxl import load_workbook

from catalog.schema import TABLES, connect, create

ROOT = Path(__file__).resolve().parents[1]
BASELINE = ROOT / "baselines/2026-09-06"
WORKBOOK_HASH = "3127e663b1531e366ce86b989b6190914108d40dfd15a33a258307a05d608e3c"
REFERENCE = "4fe92a4f078370c478f18484cad31bdafe58ad43"
NAMESPACE = uuid.UUID("e35d10b2-8928-48a6-8a04-28d51b28a52c")
EVIDENCE_FIELDS = {"notes", "note", "detail_raw", "original_detail_raw", "source", "grouping_source",
                   "Detail from Disclosure", "Color Overrides"}
ROLES = {"source_option_sheet", "status_sheet", "rule_mapping_sheet", "price_rules_sheet",
         "rule_groups_sheet", "rule_group_members_sheet", "exclusive_groups_sheet",
         "exclusive_group_members_sheet", "color_overrides_sheet", "variant_option_overrides_sheet",
         "interior_source_sheet"}


def encoded(value):
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def digest(data):
    return hashlib.sha256(data).hexdigest()


def identity(kind, model, key):
    # The namespace describes legacy identity, not a source hash or cell address.
    return str(uuid.uuid5(NAMESPACE, encoded([kind, model, key])))


def require(condition, message):
    if not condition:
        raise ValueError(message)


def boolean(value):
    if value is None:
        return None
    if value is True or value == "True":
        return 1
    if value is False or value == "False":
        return 0
    raise ValueError(f"Unrecognized boolean: {value!r}")


def integer(value):
    if value is None:
        return None
    number = Decimal(str(value))
    require(number.is_finite() and number == number.to_integral_value(), f"Invalid integer: {value!r}")
    return int(number)


def decimal(value):
    if value is None:
        return None
    try:
        number = Decimal(str(value))
    except InvalidOperation as error:
        raise ValueError(f"Invalid price: {value!r}") from error
    require(number.is_finite(), f"Nonfinite price: {value!r}")
    return format(number, "f")


def rate_type(value):
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


class Record:
    def __init__(self, sheet, ordinal, values):
        self.sheet, self.ordinal, self.values = sheet, ordinal, values
        self.id = f"{sheet}!{ordinal}"
        self.used = set()

    def get(self, field):
        require(field in self.values, f"Missing {self.sheet} header: {field}")
        self.used.add(field)
        return self.values[field]


class Importer:
    def __init__(self, db, workbook):
        self.db, self.workbook = db, workbook
        self.rows, self.expected, self.ids = {}, {}, {}
        self.models, self.roles, self.model_keys = {}, {}, {}
        self.normalizations = Counter()
        self.unresolved = []

    def take(self, row, mapping):
        return {destination: row.get(source) for destination, source in mapping.items()}

    def add(self, table, model, key, row=None, **fields):
        uid = identity(table, model, key)
        require(uid not in self.expected, f"Duplicate {table} identity: {model}/{key}")
        spec = dict(item.split(":") for item in TABLES[table][0].split())
        require(set(fields) == set(spec), f"{table} fields differ: {set(fields) ^ set(spec)}")
        for field, typ in spec.items():
            value = fields[field]
            if typ == "b":
                fields[field] = boolean(value) if not isinstance(value, int) else value
            elif typ == "i":
                fields[field] = integer(value)
            elif typ == "d":
                fields[field] = decimal(value)
            elif value is not None:
                fields[field] = str(value)
        values = {"id": uid, "model_id": model, "sequence": row.ordinal if row else 0, **fields}
        self.db.execute("INSERT INTO entity VALUES (?,?,?)", (uid, table, model))
        self.db.execute(f"INSERT INTO {table} ({','.join(values)}) VALUES ({','.join('?' for _ in values)})", tuple(values.values()))
        self.expected[uid] = (table, values)
        self.ids[(table, model, encoded(key))] = uid
        if row:
            self.link(row, uid)
        return uid

    def link(self, row, uid):
        self.db.execute("INSERT OR IGNORE INTO evidence_link VALUES (?,?)", (row.id, uid))

    def ref(self, table, model, key, optional=False):
        if optional and key is None:
            return None
        uid = self.ids.get((table, model, encoded(key)))
        require(uid is not None, f"Unresolved {table} reference: {model}/{key}")
        return uid

    def product(self, model, key):
        found = [self.ids[(kind, model, encoded(key))] for kind in ("offering", "model_interior")
                 if (kind, model, encoded(key)) in self.ids]
        require(len(found) == 1, f"Ambiguous or missing product identity: {model}/{key}")
        return found[0]

    def legacy(self, kind, model, key, uid):
        self.db.execute("INSERT INTO legacy_mapping VALUES (?,?,?,?,?)", ("27vette-v1", model, kind, key, uid))

    def disposition(self, row, model, status, reason):
        self.db.execute("INSERT OR IGNORE INTO source_disposition VALUES (?,?,?,?)", (row.id, model, status, reason))

    def capture(self):
        for sheet_index, sheet in enumerate(self.workbook, 1):
            headers = [cell.value for cell in sheet[1]]
            require(all(isinstance(h, str) and h for h in headers) and len(set(headers)) == len(headers), f"Invalid headers: {sheet.title}")
            rows = []
            for cells in sheet.iter_rows(min_row=2):
                if not any(cell.value is not None for cell in cells):
                    continue
                require(all(c.data_type != "f" for c in cells), f"Uncharacterized formula in {sheet.title}")
                values = {h: c.value for h, c in zip(headers, cells)}
                row = Record(sheet.title, cells[0].row, values)
                rows.append((row, [{"value": c.value, "data_type": c.data_type, "number_format": c.number_format} for c in cells]))
            self.db.execute("INSERT INTO source_sheet VALUES (?,?,?,?)", (sheet.title, sheet_index, encoded(headers), len(rows)))
            self.rows[sheet.title] = [r for r, _ in rows]
            for row, cells in rows:
                self.db.execute("INSERT INTO source_row VALUES (?,?,?,?)", (row.id, row.sheet, row.ordinal, encoded(cells)))

    def model_for(self, row):
        key = row.get("model_key")
        require(key in self.models, f"Unknown model: {key}")
        return self.models[key]

    def by_role(self, model, role):
        return self.rows[self.roles[model][role]]

    def scope(self, table, model, owner, row, fields):
        for field, axis in fields.items():
            raw = row.get(field)
            tokens = [] if raw is None else [s.strip() for s in raw.split("|") if s.strip()]
            # Direct-rule browser applicability is exact equality; '*' is NOT all.
            if table == "direct_rule":
                require(not tokens or tokens in (["coupe"], ["convertible"]), f"Uncharacterized direct scope at {row.id}")
            all_scope = not tokens or "*" in tokens
            sid = self.add("scope_axis", model, [owner, axis], row, owner_id=owner, axis=axis, mode="all" if all_scope else "members", all_token="*" if raw == "*" else None)
            if not all_scope:
                require(len(tokens) == len(set(tokens)), f"Duplicate scope member: {row.id}")
                for position, token in enumerate(tokens):
                    variant = self.ref("variant", model, token) if axis == "variant" else None
                    if axis == "body":
                        require(token in {"coupe", "convertible"}, f"Unknown body: {token}")
                    if axis == "trim":
                        allowed = {r[0] for r in self.db.execute("SELECT trim_level FROM variant WHERE model_id=?", (model,))}
                        require(token.lower() in {t.lower() for t in allowed}, f"Unknown trim: {token}")
                    self.add("scope_member", model, [sid, token], row, scope_id=sid, token=token, variant_id=variant, position=position)

    def load_models(self):
        for row in self.rows["model_master"]:
            key, year = row.get("model_key"), integer(row.get("model_year"))
            uid = self.add("model", "", [key, year], row, model_key=key, registry_key=row.get("registry_key"), label=row.get("model_label"), year=year,
                           active=row.get("active"), expected_variant_count=row.get("expected_variant_count"), default_model=row.get("default_model"))
            self.models[key], self.model_keys[uid] = uid, key
            source_roles = {r.get("source_role"):r.get("sheet_name") for r in self.rows["model_workbook_sources"] if r.get("model_key")==key}
            self.add("model_presentation", uid, key, row, source_workbook="stingray_master.xlsx",
                     **{f:source_roles[f] for f in ["source_option_sheet","rule_mapping_sheet","price_rules_sheet","color_overrides_sheet","interior_source_sheet"]}, **self.take(row, {f: f for f in ["dataset_name", "export_slug", "setup_card_subtitle", "setup_eyebrow", "setup_title", "setup_description"]}))
            for i in range(1, 4):
                self.add("model_fact", uid, i, row, position=i, text=row.get(f"setup_fact_{i}"))
        for row in self.rows["model_workbook_sources"]:
            model, role = self.model_for(row), row.get("source_role")
            require(role in ROLES and boolean(row.get("active")) == 1, f"Uncharacterized source assignment: {row.id}")
            require(role not in self.roles.setdefault(model, {}), f"Duplicate source assignment: {row.id}")
            self.roles[model][role] = row.get("sheet_name")
            if role in {"source_option_sheet", "rule_mapping_sheet", "price_rules_sheet", "color_overrides_sheet", "interior_source_sheet"}:
                self.link(row, self.ref("model_presentation", model, self.model_keys[model]))
            self.disposition(row, self.model_keys[model], "evidence", "Legacy workbook routing retained only as import evidence")
        require(all(set(r) == ROLES for r in self.roles.values()) and len(self.roles) == len(self.models), "Incomplete source-role registry")
        for row in self.rows["model_registry_promotion"]:
            model = self.model_for(row)
            self.add("publication", model, model, row, **self.take(row, {f:f for f in ["registry_key", "promoted_to_runtime", "default_model", "artifact_path", "artifact_type", "legacy_alias", "active", "display_order"]}))
        for row in self.rows["section_master"]:
            key = row.get("section_id")
            self.add("section", "", key, row, section_key=key, **self.take(row, {"name":"section_name", **{f:f for f in ["selection_mode", "is_required", "display_order", "standard_behavior", "step_key"]}}))
        masters = {r.get("variant_id"): r for r in self.rows["variant_master"]}
        require(len(masters) == len(self.rows["variant_master"]), "Duplicate variant master")
        for row in self.rows["model_variants"]:
            model, legacy = self.model_for(row), row.get("variant_id")
            require(legacy in masters, f"Missing variant master: {legacy}")
            master = masters[legacy]
            require(integer(master.get("model_year")) == self.db.execute("SELECT year FROM model WHERE id=?", (model,)).fetchone()[0], "Variant year mismatch")
            uid = self.add("variant", model, legacy, master, legacy_id=legacy, **self.take(master, {f:f for f in ["trim_level", "body_style", "display_name", "base_price", "display_order", "active"]}), membership_order=row.get("display_order"), membership_active=row.get("active"))
            self.link(row, uid)
            self.legacy("variant", model, legacy, uid)

    def load_options(self):
        for model in self.roles:
            for row in self.by_role(model, "source_option_sheet"):
                key = row.get("option_id")
                definition = self.add("option_definition", model, key, row, intrinsic_name=row.get("option_name"), description=row.get("description"))
                offering = self.add("offering", model, key, row, legacy_id=key, definition_id=definition, active=row.get("active"))
                self.legacy("offering", model, key, offering)
                rpo = row.get("rpo")
                if rpo is not None:
                    self.add("offering_code", model, key, row, offering_id=offering, code=rpo, role="legacy-unspecified")
                self.add("offering_price", model, key, row, offering_id=offering, amount=row.get("price"), basis="option", currency=None)
                self.add("offering_policy", model, key, row, offering_id=offering, selectable=row.get("selectable"), display_behavior=row.get("display_behavior"))
                self.add("offering_presentation", model, key, row, offering_id=offering, section_id=self.ref("section", "", row.get("section_id")), label=row.get("option_name"), description=row.get("description"), display_order=row.get("display_order"))
            for row in self.by_role(model, "status_sheet"):
                opt, var = row.get("option_id"), row.get("variant_id")
                self.add("availability", model, [opt, var], row, offering_id=self.ref("offering", model, opt), variant_id=self.ref("variant", model, var), status=row.get("status"))
            for row in self.by_role(model, "variant_option_overrides_sheet"):
                opt, var = row.get("option_id"), row.get("variant_id")
                self.add("variant_override", model, [opt, var], row, offering_id=self.ref("offering", model, opt), variant_id=self.ref("variant", model, var), section_id=self.ref("section", "", row.get("section_id"), True), **self.take(row, {f:f for f in ["selectable", "display_behavior", "active"]}))

    def load_interiors(self):
        definitions = {}
        for sheet in dict.fromkeys(r["interior_source_sheet"] for r in self.roles.values()):
            for row in self.rows[sheet]:
                key = row.get("interior_id")
                uid = self.add("interior_definition", "", key, row, legacy_id=key, source_note=row.get("Detail from Disclosure"), color_overrides_raw=row.get("Color Overrides"), section_id=self.ref("section", "", row.get("section_id")),
                    **self.take(row, {"name":"Interior Name", "material":"Material", "stored_price":"Price", "price_trim":"Trim", "seat":"Seat", "interior_code":"Interior Code", "suede":"Suede", "stitch":"Stitch", "two_tone":"Two Tone", "active_for_stingray":"active_for_stingray", "requires_r6x":"requires_r6x", "included_legacy_id":"included_option_id"}))
                definitions[key] = (uid, row)
        for row in self.rows["model_interior_scope"]:
            model, key = self.model_for(row), row.get("interior_id")
            require(key in definitions, f"Missing interior definition: {key}")
            definition, source = definitions[key]
            require(source.sheet == self.roles[model]["interior_source_sheet"], f"Interior source ownership mismatch: {row.id}")
            uid = self.add("model_interior", model, key, row, legacy_id=key, definition_id=definition, trim_level=row.get("trim_level"), active=row.get("active"),
                           requires_offering_id=self.ref("offering", model, row.get("requires_option_id"), True), included_offering_id=self.ref("offering", model, source.get("included_option_id"), True))
            self.link(source, uid)
            self.legacy("model_interior", model, key, uid)
            fields = ["seat_label", "color_family", "material_family", "variant_label", "group_display_order", "material_display_order", "choice_display_order", "parent_group_label", "leaf_label", "reference_order"]
            self.add("interior_presentation", model, key, row, interior_id=uid, **self.take(row, {f:"interior_"+f for f in fields}))
            levels = json.loads(row.get("interior_hierarchy_levels"))
            require(isinstance(levels, list) and levels and all(isinstance(s,str) and s for s in levels), f"Invalid hierarchy: {row.id}")
            parent = None
            for i, label in enumerate(levels):
                node_key = levels[:i+1]
                node_id = self.ids.get(("hierarchy_node", model, encoded(node_key)))
                if node_id is None:
                    node_id = self.add("hierarchy_node", model, node_key, row, parent_id=parent, label=label, position=i)
                else:
                    self.link(row, node_id)
                self.add("interior_hierarchy_member", model, [key, i], row, interior_id=uid, node_id=node_id, position=i)
                parent = node_id
            self.normalizations["hierarchy_arrays_to_members"] += 1
        rates = {}
        for row in self.rows["PriceRef"]:
            typ, trim, code = rate_type(row.get("OptionType")), row.get("Trim"), row.get("Code")
            trim = "" if trim is None else trim.replace("_", " ")
            key = [typ, trim, code]
            rates[tuple(key)] = self.add("component_rate", "", key, row, component_type=typ, code=code, trim_scope=trim, amount=row.get("Price"), basis="component", currency=None)
        for row in self.rows["interior_components"]:
            model, key = self.model_for(row), row.get("interior_id")
            code, typ = row.get("rpo"), row.get("component_type")
            rt, rc, trim = rate_type(row.get("price_ref_type")), row.get("price_ref_code"), row.get("price_trim_scope")
            normalized_trim = trim.replace("_", " ") if trim else ""
            rate = rates.get((rt, normalized_trim, rc), rates.get((rt, "", rc)))
            if rate is None:
                self.unresolved.append({"locator":row.id, "issue":"Missing component rate; preserve null reference, runtime fallback is D policy"})
            self.add("interior_component", model, [key, typ, code], row, interior_id=self.ref("model_interior", model, key), code=code, component_type=typ, label=row.get("label"), rate_type=rt, rate_code=rc, rate_trim=trim, rate_id=rate, display_order=row.get("display_order"), active=row.get("active"))

    def load_rules(self):
        scopes = {"body_style_scope":"body", "trim_level_scope":"trim", "variant_scope":"variant"}
        for model in self.roles:
            for row in self.by_role(model, "rule_mapping_sheet"):
                key = row.get("rule_id")
                uid = self.add("direct_rule", model, key, row, legacy_id=key, source_id=self.product(model,row.get("source_id")), target_id=self.product(model,row.get("target_id")), effect=row.get("rule_type"), runtime_action=row.get("runtime_action"), explanation=row.get("disabled_reason"), source_note=row.get("original_detail_raw"))
                self.legacy("direct_rule", model, key, uid)
                self.scope("direct_rule", model, uid, row, {"body_style_scope":"body"})
            for row in self.by_role(model, "rule_groups_sheet"):
                key = row.get("group_id")
                uid = self.add("group_rule", model, key, row, legacy_id=key, display_label=row.get("display_label"), effect=row.get("group_type"), source_id=self.product(model,row.get("source_id")), explanation=row.get("disabled_reason"), notes=row.get("notes"), active=row.get("active"))
                self.legacy("group_rule", model, key, uid)
                self.scope("group_rule", model, uid, row, scopes)
            for row in self.by_role(model, "rule_group_members_sheet"):
                group, target = row.get("group_id"), row.get("target_id")
                self.add("group_member", model, [group,target], row, group_id=self.ref("group_rule",model,group), target_id=self.product(model,target), display_order=row.get("display_order"), active=row.get("active"))
            for row in self.by_role(model, "exclusive_groups_sheet"):
                key, mode = row.get("group_id"), row.get("selection_mode")
                modes = {"single_within_group":"at_most_one", "required_single_within_group":"exactly_one"}
                require(mode in modes, f"Unknown exclusive selection mode: {mode}")
                uid = self.add("exclusive_group", model, key, row, legacy_id=key, display_label=row.get("display_label"), selection_mode=modes[mode], notes=row.get("notes"), active=row.get("active"))
                self.legacy("exclusive_group", model, key, uid)
            for row in self.by_role(model, "exclusive_group_members_sheet"):
                group, option = row.get("group_id"), row.get("option_id")
                self.add("exclusive_member", model, [group,option], row, group_id=self.ref("exclusive_group",model,group), offering_id=self.ref("offering",model,option), display_order=row.get("display_order"), active=row.get("active"))
            for row in self.by_role(model, "price_rules_sheet"):
                key = row.get("price_rule_id")
                uid = self.add("price_rule", model, key, row, legacy_id=key, condition_id=self.product(model,row.get("condition_option_id")), target_id=self.ref("offering",model,row.get("target_option_id")), effect=row.get("price_rule_type"), amount=row.get("price_value"), basis="conditional_total", currency=None, notes=row.get("notes"))
                self.legacy("price_rule", model, key, uid)
                self.scope("price_rule", model, uid, row, {k:v for k,v in scopes.items() if k != "variant_scope"})
            for row in self.by_role(model, "color_overrides_sheet"):
                key = row.get("interior_id")
                if ("model_interior", model, encoded(key)) not in self.ids:
                    self.disposition(row, self.model_keys[model], "not_applicable", "Interior absent from this model's explicit interior membership; inspection.py:build_color_overrides")
                    continue
                condition, added = row.get("option_id"), row.get("adds_rpo")
                self.add("color_rule", model, [key,condition,added], row, interior_id=self.ref("model_interior",model,key), condition_id=self.ref("offering",model,condition), added_id=self.ref("offering",model,added), effect=row.get("rule_type"))
        for row in self.rows["default_selection_rules"]:
            model, key = self.model_for(row), row.get("rule_id")
            kind, operand = row.get("condition_type"), row.get("condition_id")
            require(kind in {"always","unless_selected_rpo","unless_selected_section","when_selected_unless_selected_section"}, f"Unknown default condition: {kind}")
            require(kind != "always" or operand is None, f"Unexpected always operand: {row.id}")
            uid = self.add("default_rule", model, key, row, legacy_id=key, target_id=self.ref("offering",model,row.get("target_option_id")), condition_kind=kind,
                condition_code=operand if kind == "unless_selected_rpo" else None,
                condition_section_id=self.ref("section","",operand) if kind == "unless_selected_section" else None,
                condition_offering_id=self.ref("offering",model,operand) if kind == "when_selected_unless_selected_section" else None,
                target_section_mode="resolved_target_section" if kind == "when_selected_unless_selected_section" else None,
                notes=row.get("notes"), priority=row.get("priority"), display_behavior=row.get("display_behavior"), active=row.get("active"))
            self.legacy("default_rule", model, key, uid)
            self.scope("default_rule", model, uid, row, scopes)

    def load_presentation(self):
        for row in self.rows["context_section_master"]:
            model, key = self.model_for(row), row.get("context_type")
            self.add("context_section",model,key,row,context_type=key, **self.take(row,{"section_key":"section_id","name":"section_name","display_order":"section_display_order",**{f:f for f in ["selection_mode","choice_mode","is_required","standard_behavior","step_key","step_label","active"]}}))
        for row in self.rows["section_presentation"]:
            model, key = self.model_for(row), row.get("section_id")
            self.add("section_presentation",model,key,row,section_id=self.ref("section","",key), **self.take(row,{"display_order":"section_display_order",**{f:f for f in ["display_label","step_key","display_behavior","standard_equipment_bucket","standard_equipment_group_type","auto_added_bucket","active"]}}))
        for row in self.rows["runtime_steps"]:
            model, key = self.model_for(row), row.get("step_key")
            self.add("runtime_step",model,key,row,step_key=key,label=row.get("step_label"),runtime_order=row.get("runtime_order"),active=row.get("active"),navigable=True,source=row.get("source"))
        for row in self.rows["order_summary_sections"]:
            model, key = self.model_for(row), row.get("section_key")
            self.add("summary_section",model,key,row,section_key=key,notes=row.get("notes"),label=row.get("section_label"),display_order=row.get("display_order"),active=row.get("active"))
        for row in self.rows["step_order_summary_map"]:
            model, step = self.model_for(row), row.get("step_key")
            if ("runtime_step", model, encoded(step)) not in self.ids:
                require(step == "standard_equipment", f"Unknown summary step: {row.id}")
                self.add("runtime_step",model,step,row,step_key=step,label="Standard Equipment",runtime_order=None,active=True,navigable=False,source=None)
                self.normalizations["code_owned_nonnavigable_bucket"] += 1
            self.add("step_summary",model,step,row,step_id=self.ref("runtime_step",model,step),section_id=self.ref("summary_section",model,row.get("section_key")),active=row.get("active"))
        # Context cards are explicit combinations derived only from imported variants.
        for model in self.roles:
            for column, kind in [("body_style","body_style"),("trim_level","trim_level")]:
                for (value,) in self.db.execute(f"SELECT DISTINCT {column} FROM variant WHERE model_id=? ORDER BY {column}",(model,)).fetchall():
                    self.add("context_choice",model,[kind,value],context_type=kind,value=value)
        for row in self.rows["context_choice_copy"]:
            model_key, kind, value, body = row.get("model_key"),row.get("context_type"),row.get("value"),row.get("body_style")
            require(model_key == "*" or model_key in self.models, f"Invalid copy model: {row.id}")
            for key, model in self.models.items():
                if model_key not in ("*",key):
                    continue
                self.add("context_copy",model,[kind,value,body],row,context_type=kind,value=value,body_scope=body,info_tooltip=row.get("info_tooltip"),active=row.get("active"))
        for row in self.rows["asset_map"]:
            model_key, kind, key = row.get("model_key"),row.get("target_type"),row.get("target_id")
            require(model_key == "*" or model_key in self.models, f"Unknown asset model: {row.id}")
            for name, model in self.models.items():
                if model_key not in ("*",name):
                    continue
                if model_key == "*" and kind == "option" and ("offering",model,encoded(key)) not in self.ids:
                    self.disposition(row,name,"not_applicable","Shared asset target absent from model offerings")
                    continue
                self.load_asset(row, model, model_key, kind, key)

    def load_asset(self, row, model, model_key, kind, key):
        require(model_key != "*" or kind == "option", f"Unsupported shared asset type: {kind}")
        if kind == "option":
            target = self.ref("offering",model,key)
        elif kind == "context_choice":
            target = self.ref("context_choice",model,key.split("__",1))
        elif kind == "model":
            require(key == self.db.execute("SELECT registry_key FROM model WHERE id=?",(model,)).fetchone()[0], f"Model asset identity mismatch: {row.id}")
            # Model card assets attach to the model-scoped presentation.
            target = self.ref("model_presentation",model,self.model_keys[model])
        else:
            raise ValueError(f"Unsupported asset target: {kind}")
        self.add("asset_assignment",model,[model_key,kind,key],row,target_id=target,source_scope="shared" if model_key=="*" else "model",**self.take(row,{f:f for f in ["image_url","image_alt","image_fit","image_position","hover_image_url","hover_image_alt","hover_image_position","active"]}))

    def code_policies(self):
        policies = [
            ("scripts/corvette_form_generator/rule_derivation.py","EMISSION_ALLOWLIST","Five Z06 approved includes-closure replacement pairs; typed permission rows retained"),
            ("scripts/corvette_form_generator/pricing.py","generated_interior_price","Stored interior amount plus max(0,R6X seat rate minus base seat rate); missing lookup falls back to zero"),
            ("scripts/corvette_form_generator/pricing.py","price_ref_component_price","Normalize type by removing non-alphanumeric lowercase characters, replace trim underscores with spaces, exact trim then universal fallback"),
            ("scripts/corvette_form_generator/interiors.py","build_model_interiors","Explicit model scope controls membership; R6X runtime flag inferred from trim/ID independently of authored requires_r6x"),
            ("scripts/corvette_form_generator/rules.py","build_draft_rules","Preserve all authored rows; runtime removes grouped requirements and disables redundant same-section exclusions; explanations do not deactivate source rows"),
            ("scripts/corvette_form_generator/inspection.py","build_color_overrides","Filter shared color rows by model interior membership before resolving offering references"),
            ("scripts/corvette_form_generator/runtime_metadata.py","BUCKET_STEP_KEYS / UNAUTHORED_BUCKET_STEP_LABELS","standard_equipment is a nonnavigable bucket, absent from runtime_steps; label is Standard Equipment"),
            ("scripts/corvette_form_generator/contract.py","load_asset_map","Active explicit-model asset rows override shared option assets; retain both source scopes"),
            ("scripts/corvette_form_generator/inspection.py","build_draft_price_rules","Conditional price source may be an offering or model interior; target must be an offering"),
            ("form-app/app.js","optionPrice","First matching conditional override; package component base is minimum positive qualifying price, selected component delta has zero floor"),
            ("form-app/app.js","addGeneratedDefaultChoices","When-selected defaults test target section resolved for current variant, including presentation overrides"),
            ("form-app/app.js","scopeMatches","Pipe-delimited exact case-sensitive membership; blank or star matches all; direct rules use different exact body check"),
        ]
        self.db.executemany("INSERT INTO code_evidence VALUES (?,?,?,?,?)",[(p,REFERENCE,s,"baseline_behavior_pending_D",d) for p,s,d in policies])
        model = self.models["z06"]
        for code in ["t0f","t0g","z07","pdd","pdf"]:
            source, target = f"opt_{code}_001", "opt_cbf_001"
            self.add("derivation_permission",model,[source,target],source_id=self.ref("offering",model,source),target_id=self.ref("offering",model,target),method="includes_closure_approved_replace")

    def reconcile(self):
        # Independently read back every typed row/field and its order. The expected
        # values are formed from workbook rows before SQLite receives them.
        for uid,(table,expected) in self.expected.items():
            actual = self.db.execute(f"SELECT * FROM {table} WHERE id=?",(uid,)).fetchone()
            require(actual is not None and dict(actual) == expected, f"Typed value reconciliation failed: {table}/{uid}")
        counts = {t:self.db.execute(f"SELECT count(*) FROM {t}").fetchone()[0] for t in TABLES}
        require(sum(counts.values()) == len(self.expected), "Unexpected typed rows")
        registry = {r["id"]:(r["kind"],r["model_id"]) for r in self.db.execute("SELECT * FROM entity")}
        require(registry == {uid:(t,v["model_id"]) for uid,(t,v) in self.expected.items()}, "Entity registry/subtype mismatch")
        require(not self.db.execute("PRAGMA foreign_key_check").fetchall(), "Foreign key violation")
        require(self.db.execute("PRAGMA integrity_check").fetchone()[0] == "ok", "SQLite integrity failure")
        for model,key in self.model_keys.items():
            variants = self.db.execute("SELECT count(*) FROM variant WHERE model_id=?",(model,)).fetchone()[0]
            expected = self.db.execute("SELECT expected_variant_count FROM model WHERE id=?",(model,)).fetchone()[0]
            require(variants == expected, f"Variant coverage mismatch: {key}")
            require(not self.db.execute("""SELECT trim_level,body_style FROM variant WHERE model_id=?
                GROUP BY trim_level,body_style HAVING count(*)>1""",(model,)).fetchall(), f"Variant natural key collision: {key}")
            # Cartesian equality catches missing memberships, not only total counts.
            missing = self.db.execute("""SELECT o.id,v.id FROM offering o JOIN variant v ON o.model_id=v.model_id
                LEFT JOIN availability a ON a.offering_id=o.id AND a.variant_id=v.id
                WHERE o.model_id=? AND a.id IS NULL""",(model,)).fetchall()
            require(not missing, f"Missing availability pairs: {key}")
        require(self.db.execute("SELECT count(*) FROM publication WHERE active=1 AND promoted_to_runtime=1 AND default_model=1").fetchone()[0] == 1, "Publication must have one default")
        require(not self.db.execute("""SELECT p.id FROM publication p JOIN model m ON m.id=p.model_id
            WHERE p.default_model != m.default_model OR p.registry_key != m.registry_key""").fetchall(), "Conflicting publication metadata")
        require(not self.db.execute("""SELECT s.id FROM scope_axis s LEFT JOIN scope_member m ON m.scope_id=s.id
            GROUP BY s.id HAVING (s.mode='all' AND count(m.id)>0) OR (s.mode='members' AND count(m.id)=0)""").fetchall(), "Invalid scope membership")
        accounting = []
        for sheet,rows in self.rows.items():
            for row in rows:
                if sheet == "rule_phrase_map":
                    row.used.update(row.values)
                    self.disposition(row,"","evidence","Historical parser phrase map; no executable authority")
                unhandled = set(row.values) - row.used - EVIDENCE_FIELDS
                require(not unhandled, f"Unmapped fields at {row.id}: {sorted(unhandled)}")
                linked = self.db.execute("SELECT count(*) FROM evidence_link WHERE source_id=?",(row.id,)).fetchone()[0]
                explained = self.db.execute("SELECT count(*) FROM source_disposition WHERE source_id=?",(row.id,)).fetchone()[0]
                require(linked or explained, f"Unaccounted source row: {row.id}")
            accounting.append({"sheet":sheet,"source_rows":len(rows),"accounted_rows":len(rows)})
        models = {key:{table:self.db.execute(f"SELECT count(*) FROM {table} WHERE model_id=?",(model,)).fetchone()[0]
                       for table in ("variant","offering","availability","model_interior","direct_rule","group_rule","price_rule")}
                  for model,key in self.model_keys.items()}
        dispositions = [dict(r) for r in self.db.execute("SELECT model_key,disposition,reason,count(*) AS rows FROM source_disposition GROUP BY model_key,disposition,reason ORDER BY model_key,disposition,reason")]
        return {"source_sheets":accounting,"models":models,"typed_rows":counts,"reconciled_fields":sum(len(v) for _,v in self.expected.values()),
                "source_dispositions":dispositions,
                "added_business_facts":[],"changed_business_facts":[],"removed_business_facts":[],
                "normalizations":dict(self.normalizations),"unresolved":self.unresolved,
                "limitations":["Currency remains unknown; exact source amounts retained", "Runtime policies recorded with pinned code provenance; execution and six-model contract parity belong to D", "Disposable candidate schema; authoring/revision/release operations are not implemented"]}

    def run(self):
        self.capture()
        require(len(self.rows)==77, "Expected the inspected 77-sheet baseline")
        require(not self.rows["runtime_rule_exceptions"], "Populated runtime exceptions require explicit classification")
        self.load_models()
        self.load_options()
        self.load_interiors()
        self.load_rules()
        self.load_presentation()
        self.code_policies()
        return self.reconcile()


def workbook_bytes(baseline=BASELINE):
    baseline = Path(baseline)
    manifest = json.loads((baseline/"manifest.json").read_text())
    archive = (baseline/manifest["archive"]["path"]).read_bytes()
    require(digest(archive)==manifest["archive"]["sha256"], "Baseline archive hash mismatch")
    with tarfile.open(fileobj=io.BytesIO(archive),mode="r:gz") as tar:
        data=tar.extractfile("stingray_master.xlsx").read()
    require(digest(data)==WORKBOOK_HASH, "Uninspected workbook hash; register a new baseline before importing")
    return data


def build(output, baseline=BASELINE):
    """Validate a temporary sibling DB, then publish without overwriting anything."""
    data=workbook_bytes(baseline)
    output=Path(output)
    output.parent.mkdir(parents=True,exist_ok=True)
    fd,tmp=tempfile.mkstemp(prefix=".candidate-",suffix=".sqlite",dir=output.parent)
    os.close(fd)
    db = None
    wb = None
    try:
        db=connect(tmp)
        wb=load_workbook(io.BytesIO(data),data_only=False)
        create(db)
        with db:
            db.executemany("INSERT INTO import_metadata VALUES (?,?)",[("workbook_sha256",WORKBOOK_HASH),("reference_commit",REFERENCE),("schema_version","2"),("authority","disposable_candidate"),
                ("importer_sha256",digest(Path(__file__).read_bytes())),("schema_sha256",digest((ROOT/"catalog/schema.py").read_bytes()))])
            report=Importer(db,wb).run()
            db.execute("INSERT INTO import_metadata VALUES (?,?)",("reconciliation",encoded(report)))
        db.close()
        # Hard link publishes a fully committed DB atomically without replacing
        # an existing destination, including one another process just created.
        try:
            os.link(tmp,output)
        except FileExistsError:
            require(Path(tmp).read_bytes() == output.read_bytes(), f"Output differs; refusing to overwrite: {output}")
        return report
    finally:
        if wb is not None:
            wb.close()
        if db is not None:
            db.close()
        Path(tmp).unlink(missing_ok=True)


def main():
    parser=argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output",type=Path,required=True)
    parser.add_argument("--baseline",type=Path,default=BASELINE)
    args=parser.parse_args()
    report=build(args.output,args.baseline)
    print(json.dumps({"output":str(args.output),"source_rows":sum(s["source_rows"] for s in report["source_sheets"]),"typed_rows":report["typed_rows"],"unresolved":report["unresolved"]},indent=2))


if __name__=="__main__":
    main()
