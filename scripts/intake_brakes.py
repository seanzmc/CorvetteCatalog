"""Bounded MY2027 Stingray brake intake. Writes immutable review JSON only."""
from __future__ import annotations

import argparse
from collections import Counter
from decimal import Decimal
import hashlib
import io
import json
import os
from pathlib import Path
import re
import tarfile
import tempfile
from xml.etree.ElementTree import tostring

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
BASELINE = ROOT / "baselines/2026-09-06"
SOURCE_HASH = "d3ca7d3a09c9fb89210b4ce584493b3ad8fb65ca35087c49d816d1cbf1a333d1"
SOURCE = ROOT / "sources/raw" / SOURCE_HASH / "2027 Chevrolet Car Corvette Export (6).xlsx"
CONFIG = {"family": "MY2027 Stingray JL9/J55 and Z51", "version": 1,
          "mechanical_ranges": ["A1:I3", "A6:I7", "A53:I53"],
          "price_ranges": ["A2:J4", "B7:J10", "B44:J46", "B229:J229", "A298:A308"],
          "normalization": "S/A/-- with separately resolved superscript 1; no price acceptance"}


def encoded(value):
    return (json.dumps(value, ensure_ascii=False, indent=2) + "\n").encode()


def digest(value):
    return hashlib.sha256(value).hexdigest()


def require(condition, message):
    if not condition:
        raise ValueError(message)


def cell_evidence(cell):
    value = cell.value
    runs = []
    if isinstance(value, list):
        for part in value:
            font = getattr(part, "font", None)
            runs.append({"text": str(part), "vertical_align": getattr(font, "vertAlign", None),
                         "font_xml": tostring(font.to_tree(), encoding="unicode") if font else None})
        value = str(value)
    return {"value": value, "data_type": cell.data_type, "runs": runs,
            "number_format": cell.number_format}


def availability(cell):
    evidence = cell_evidence(cell)
    runs = evidence["runs"]
    symbol = "".join(r["text"] for r in runs if r["vertical_align"] != "superscript") if runs else str(cell.value)
    notes = [r["text"] for r in runs if r["vertical_align"] == "superscript"]
    issues = []
    if symbol not in {"S", "A", "--"}:
        issues.append("Unrecognized availability symbol; never infer a footnote from flattened A1.")
    if notes not in ([], ["1"]):
        issues.append("Unresolved footnote marker.")
    return {"S": "standard", "A": "available", "--": "unavailable"}.get(symbol), notes, issues


def records(sheet):
    headers = [c.value for c in sheet[1]]
    return [(i, dict(zip(headers, values))) for i, values in
            enumerate(sheet.iter_rows(min_row=2, values_only=True), 2) if any(v is not None for v in values)]


def only(items, label):
    require(len(items) == 1, f"Expected one {label}; found {len(items)}")
    return items[0]


def build(source=SOURCE, baseline=BASELINE):
    source = Path(source)
    raw = source.read_bytes()
    # This first adapter deliberately supports only the inspected document.
    require(digest(raw) == SOURCE_HASH, "Uninspected source revision: register and inspect before extending this pilot")
    manifest = json.loads((baseline / "manifest.json").read_text())
    archive = (baseline / manifest["archive"]["path"]).read_bytes()
    require(digest(archive) == manifest["archive"]["sha256"], "Baseline archive hash mismatch")
    with tarfile.open(fileobj=io.BytesIO(archive), mode="r:gz") as tar:
        workbook_bytes = tar.extractfile("stingray_master.xlsx").read()
        runtime_bytes = tar.extractfile("form-output/runtime/stingray-runtime-contract.json").read()
    for name, data in [("stingray_master.xlsx", workbook_bytes),
                       ("form-output/runtime/stingray-runtime-contract.json", runtime_bytes)]:
        expected = only([e for e in manifest["files"] if e["path"] == name], name)
        require(digest(data) == expected["sha256"], f"Baseline member hash mismatch: {name}")
    guide = load_workbook(io.BytesIO(raw), rich_text=True, data_only=False)
    workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=False)
    try:
        return extract(guide, workbook, json.loads(runtime_bytes), digest(workbook_bytes), digest(runtime_bytes))
    finally:
        guide.close()
        workbook.close()


def extract(guide, workbook, runtime, baseline_hash, runtime_hash):
    mechanical = guide["Mechanical 1"]
    price = guide["Price Schedule"]
    require(mechanical["A1"].value == "Stingray", "Wrong model")
    require(str(price["A2"].value) == "2027 CHEVROLET CORVETTE", "Wrong model year")
    for ref, expected in {"A3": "Orderable RPO Code", "B3": "Ref. Only RPO Code",
                          "B6": "JL9", "B7": "J55", "A53": "Z51"}.items():
        require(str(mechanical[ref].value) == expected, f"Unexpected mechanical layout at {ref}")
    require(price["B229"].value == "Z51" and price["B44"].value == "Option Code", "Unexpected price layout")
    evidence, coverage = {}, []
    selected = {"Mechanical 1": CONFIG["mechanical_ranges"], "Price Schedule": CONFIG["price_ranges"]}
    for sheet_name, ranges in selected.items():
        sheet = guide[sheet_name]
        for region in ranges:
            cells = [cell for row in sheet[region] for cell in row]
            for cell in cells:
                evidence[f"{sheet_name}!{cell.coordinate}"] = cell_evidence(cell)
            coverage.append({"locator": f"{sheet_name}!{region}", "status": "partial",
                             "cells": len(cells), "nonblank_cells": sum(c.value is not None for c in cells),
                             "reason": "Exact cells/runs retained; only listed candidate assertions interpreted. Prose and unrelated package members remain context."})
    inventory = []
    for sheet in guide:
        inventory.append({"sheet": sheet.title, "status": "partial" if sheet.title in selected else "out_of_scope",
                          "populated_rows": sum(any(c.value is not None for c in row) for row in sheet),
                          "sheet_state": sheet.sheet_state,
                          "merged_ranges": sorted(str(r) for r in sheet.merged_cells.ranges),
                          "hidden_rows": [i for i, d in sheet.row_dimensions.items() if d.hidden],
                          "hidden_columns": [i for i, d in sheet.column_dimensions.items() if d.hidden]})
    parser_hash = digest(Path(__file__).read_bytes())
    run_id = digest(encoded([SOURCE_HASH, parser_hash, CONFIG]))
    comparison_id = digest(encoded([run_id, baseline_hash, runtime_hash]))
    candidates, mappings = [], []
    option_rows = records(workbook["stingray_options"])
    options = {}
    # Mappings use explicit inspected legacy IDs, with RPO/active assertions;
    # they never merge offerings solely because a code or label matches.
    for code, legacy in [("JL9", "opt_jl9_001"), ("J55", "opt_j55_001"), ("Z51", "opt_z51_001")]:
        row, record = only([(i, r) for i, r in option_rows if r["option_id"] == legacy], legacy)
        require(record["rpo"] == code and record["active"] is True, f"Legacy mapping changed: {legacy}")
        options[code] = (row, record)
        mappings.append({"model": "stingray", "year": 2027, "source_code": code,
                         "target_legacy_id": legacy, "status": "proposed",
                         "basis": f"Inspected Checkpoint A mapping; verified stingray_options!A{row}:K{row}; no shared identity inferred."})
    statuses = records(workbook["stingray_ovs"])
    variant_rows = records(workbook["variant_master"])
    memberships = [(i, r) for i, r in records(workbook["model_variants"])
                   if r["model_key"] == "stingray" and r["active"] is True]
    member_ids = [r["variant_id"] for _, r in memberships]
    require(len(member_ids) == len(set(member_ids)), "Duplicate active Stingray membership")
    variants, membership_refs = [], {}
    for membership_row, membership in memberships:
        variant_id = membership["variant_id"]
        row, variant = only([(i, r) for i, r in variant_rows if r["variant_id"] == variant_id],
                            f"variant identity {variant_id}")
        require(variant["active"] is True, f"Inactive variant in active membership: {variant_id}")
        if variant["model_year"] == 2027:
            variants.append((row, variant))
            membership_refs[variant_id] = f"model_variants!A{membership_row}:E{membership_row}"
    all_variants, variant_mappings = [], []

    def candidate(kind, value, current, source_refs, baseline_refs, issues=(), scope=None, reason=""):
        ordinal = len(candidates)
        candidates.append({"id": digest(encoded([run_id, source_refs, kind, ordinal])),
                           "kind": kind, "typed_value": value, "baseline_value": current,
                           "applicability": scope or {"model": "stingray", "year": 2027, "variant_ids": all_variants.copy()},
                           "source_evidence": source_refs, "baseline_evidence": baseline_refs,
                           "parse_issues": list(issues), "comparison_class": "ambiguous" if issues else
                           ("unchanged" if value == current else "changed"),
                           "reason": reason, "status": "pending_review"})

    note = "1. Included and only available with (Z51) Z51 Performance Package."
    note_valid = str(mechanical["C7"].value).split("\n")[-1] == note
    for column in "DEFGHI":
        heading = str(mechanical[f"{column}3"].value)
        match = re.fullmatch(r"Stingray (Coupe|Convertible)\n(1YC07|1YC67)\n([123]LT)", heading)
        require(match is not None, f"Unrecognized variant heading: {heading}")
        body, model_code, trim = match.groups()
        require(model_code == ("1YC07" if body == "Coupe" else "1YC67"), "Inconsistent variant body/code")
        row, variant = only([(i, r) for i, r in variants if r["trim_level"] == trim.lower() and
                             r["body_style"] == body.lower() and r["model_year"] == 2027], heading)
        variant_id = variant["variant_id"]
        require(variant_id not in all_variants, "Duplicate variant identity")
        all_variants.append(variant_id)
        variant_refs = [f"variant_master!A{row}:H{row}", membership_refs[variant_id]]
        variant_mappings.append({"source_heading": f"Mechanical 1!{column}3",
                                 "model": "stingray", "year": 2027, "trim": trim,
                                 "body": body.lower(), "manufacturer_model_code": model_code,
                                 "target_variant_id": variant_id, "status": "proposed",
                                 "baseline_evidence": variant_refs})
        for code, source_row in [("JL9", 6), ("J55", 7), ("Z51", 53)]:
            source_cell = mechanical[f"{column}{source_row}"]
            status, notes, issues = availability(source_cell)
            if notes and (code != "J55" or not note_valid):
                issues.append("Footnote cannot be resolved within the pilot")
            if code == "J55" and notes != ["1"]:
                issues.append("Expected package restriction footnote is missing")
            legacy = options[code][1]["option_id"]
            baseline_row, baseline_record = only([(i, r) for i, r in statuses if
                                                  r["option_id"] == legacy and r["variant_id"] == variant_id], "availability pair")
            candidate("availability", status, baseline_record["status"],
                      [f"Mechanical 1!{column}{source_row}", f"Mechanical 1!{column}3", "Mechanical 1!C2"] +
                      (["Mechanical 1!C7"] if notes else []),
                      [f"stingray_ovs!A{baseline_row}:C{baseline_row}"] + variant_refs, issues,
                      {"model": "stingray", "year": 2027, "variant_id": variant_id, "legacy_option_id": legacy,
                       "manufacturer_model_code": model_code, "trim": trim, "body": body.lower()})
    require(set(all_variants) == {r["variant_id"] for _, r in variants},
            "Guide headings do not cover all active MY2027 Stingray variants")
    rules = records(workbook["rule_mapping"])
    include_rows = [(i, r) for i, r in rules if r["source_id"] == "opt_z51_001" and
                    r["target_id"] == "opt_j55_001" and r["rule_type"] == "includes"]
    included = "includes (J55)" in str(mechanical["C53"].value) and note_valid
    require_rows = [(i, r) for i, r in rules if r["source_id"] == "opt_j55_001" and
                    r["target_id"] == "opt_z51_001" and r["rule_type"] == "requires"]
    candidate("includes", {"source": "Z51", "target": "J55"} if included else None,
              {"source": "Z51", "target": "J55"} if len(include_rows) == 1 else None,
              ["Mechanical 1!C53", "Mechanical 1!C7"], [f"rule_mapping!A{i}:H{i}" for i, _ in include_rows],
              [] if included and len(include_rows) == 1 else ["Unresolved inclusion wording or baseline mapping"])
    candidate("requires", {"source": "J55", "target": "Z51"} if note_valid else None,
              {"explicit_rule_count": len(require_rows), "selectable": options["J55"][1]["selectable"]},
              ["Mechanical 1!C7"], [f"stingray_options!A{options['J55'][0]}:K{options['J55'][0]}",
                                      f"rule_mapping!A1:H{workbook['rule_mapping'].max_row}"],
              ["No direct prerequisite rule in baseline; nonselectability plus inclusion may enforce it. Behavioral equivalence remains unverified."],
              reason="Representation gap, not an accepted added rule.")
    candidate("replaces", {"selected": "Z51", "removed": "JL9", "included": "J55"}, None,
              ["Mechanical 1!C6", "Mechanical 1!C7", "Mechanical 1!C53"],
              [f"rule_mapping!A1:H{workbook['rule_mapping'].max_row}"],
              ["Owner-confirmed interpretation in docs/source-schema-specification.md section 8; guide cells do not explicitly state removal of JL9. No matching rule in inspected baseline."],
              reason="Keep owner interpretation separate from extracted manufacturer wording; no automatic correction.")
    for code, ref, role in [("JL9", "B6", "reference"), ("J55", "B7", "reference"), ("Z51", "A53", "orderable")]:
        row, record = options[code]
        candidate("code_role", {"code": code, "role": role}, {"rpo": record["rpo"], "role": None},
                  [f"Mechanical 1!{ref}", "Mechanical 1!A3", "Mechanical 1!B3"], [f"stingray_options!B{row}"],
                  ["Baseline RPO field does not encode orderable versus reference-only role."])
    candidate("option_price", {"header_labeled_msrp": str(price["F229"].value),
                               "apparent_shifted_amount": str(price["G229"].value), "currency": None},
              {"amount": str(options["Z51"][1]["price"]), "basis": "option"},
              ["Price Schedule!F44", "Price Schedule!G44", "Price Schedule!D229", "Price Schedule!F229", "Price Schedule!G229"],
              [f"stingray_options!C{options['Z51'][0]}"], ["Option headers appear shifted: F header says MSRP but F229=0; G229=5395. Currency is not established."],
              reason="5395 numerically matches the workbook, but header/basis ambiguity prevents price equality or a proposed zero-price correction.")
    total = Decimal(str(price["F10"].value)) + Decimal(str(price["J10"].value))
    row, base = only([(i, r) for i, r in variants if r["variant_id"] == "1lt_c07"], "base variant")
    candidate("base_price_context", {"amount": str(total), "basis": "baseline_total", "currency": None},
              {"amount": str(base["base_price"]), "basis": "baseline_total", "currency": None},
              ["Price Schedule!B10", "Price Schedule!C10", "Price Schedule!F7", "Price Schedule!J7", "Price Schedule!F10", "Price Schedule!J10", "Price Schedule!A304"],
              [f"variant_master!A{row}:H{row}", membership_refs[base["variant_id"]]],
              ["Currency remains unconfirmed; numeric MSRP plus destination comparison only."],
              scope={"model": "stingray", "year": 2027, "variant_id": base["variant_id"]})
    package_refs = sorted(set(re.findall(r"\(([A-Z0-9]{3})\)", str(mechanical["C53"].value))) - {"J55"})
    revision = str(price["A308"].value)
    require("Revised July 06, 2026" in revision, "Unexpected revision label")
    return {"run_id": run_id, "comparison_id": comparison_id, "parser_sha256": parser_hash,
            "config": CONFIG, "source_sha256": SOURCE_HASH, "baseline_workbook_sha256": baseline_hash,
            "baseline_runtime_sha256": runtime_hash, "status": "pending_review", "accepted_operations": [],
            "source_metadata": {"received_date": "2026-09-06", "received_date_timezone": "UTC",
                                "acquisition_history": "User supplied local file; earlier download history unknown",
                                "price_schedule_revision": "2026-07-06", "revision_raw": revision,
                                "revision_locator": "Price Schedule!A308", "whole_guide_revision": None,
                                "currency": None},
            "inventory": inventory, "coverage": coverage, "evidence": evidence, "mappings": mappings,
            "variant_mappings": variant_mappings,
            "candidates": candidates, "comparison_counts": dict(sorted(Counter(c["comparison_class"] for c in candidates).items())),
            "unresolved": {"external_package_codes": package_refs,
                           "unparsed_text": "Descriptions, dimensional claims, legend symbols outside S/A/-- and unrelated price notes retained without full interpretation.",
                           "removal_authority": False, "whole_guide_coverage": False},
            "runtime_evidence": [{"locator": f"/rules/{i}", "value": r} for i, r in enumerate(runtime["rules"])
                                 if r.get("source_id") in {"opt_j55_001", "opt_jl9_001"} or r.get("target_id") in {"opt_j55_001", "opt_jl9_001"}]}


def save(result, output):
    output = Path(output)
    output.mkdir(parents=True, exist_ok=True)
    target = output / (result["comparison_id"] + ".json")
    content = encoded(result)
    if target.exists():
        require(target.read_bytes() == content, "Existing staging content differs; refusing overwrite")
        return target, "reused"
    # Publish a complete file without replacing an existing comparison.
    with tempfile.NamedTemporaryFile(dir=output, delete=False) as handle:
        temporary = Path(handle.name)
        handle.write(content)
    try:
        try:
            os.link(temporary, target)
        except FileExistsError:
            require(target.read_bytes() == content, "Existing staging content differs; refusing overwrite")
            return target, "reused"
    finally:
        temporary.unlink()
    return target, "created"


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=SOURCE)
    parser.add_argument("--output", type=Path, default=ROOT / "intake/stingray-brakes")
    args = parser.parse_args()
    result = build(args.source)
    target, status = save(result, args.output)
    print(json.dumps({"path": str(target), "status": status, "counts": result["comparison_counts"]}))
