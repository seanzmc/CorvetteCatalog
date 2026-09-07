"""Frozen-source reconciliation plus corrupt-input and DB-boundary regressions."""
import io
import json
from pathlib import Path
import sqlite3
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import load_workbook

from catalog.importer import Importer, build, workbook_bytes
from catalog.schema import connect, create


class BaselineImporterTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.temp = tempfile.TemporaryDirectory()
        cls.path = Path(cls.temp.name)/"candidate.sqlite"
        cls.report = build(cls.path)
        cls.raw = workbook_bytes()
        cls.wb = load_workbook(io.BytesIO(cls.raw))

    @classmethod
    def tearDownClass(cls):
        cls.wb.close()
        cls.temp.cleanup()

    def setUp(self):
        self.db = connect(self.path)
        self.db.execute("SAVEPOINT test")

    def tearDown(self):
        self.db.execute("ROLLBACK TO test")
        self.db.close()

    def rows(self, name):
        rows = list(self.wb[name].values)
        return [dict(zip(rows[0],r)) for r in rows[1:] if any(v is not None for v in r)]

    def test_all_source_cells_types_and_order_preserved(self):
        self.assertEqual(sum(s["source_rows"] for s in self.report["source_sheets"]),15134)
        self.assertEqual(len(self.report["source_sheets"]),77)
        for sheet in self.wb:
            actual = {r["ordinal"]:json.loads(r["cells"]) for r in self.db.execute("SELECT ordinal,cells FROM source_row WHERE sheet=?",(sheet.title,))}
            expected = {r[0].row:[{"value":c.value,"data_type":c.data_type,"number_format":c.number_format} for c in r]
                        for r in sheet.iter_rows(min_row=2) if any(c.value is not None for c in r)}
            self.assertEqual(actual,expected,sheet.title)

    def test_six_model_options_prices_presentation_and_availability(self):
        for model in self.db.execute("SELECT id,model_key FROM model"):
            mid,key = model
            roles = {r["source_role"]:r["sheet_name"] for r in self.rows("model_workbook_sources") if r["model_key"]==key}
            actual = [tuple(r) for r in self.db.execute("""SELECT o.legacy_id,o.rpo,o.base_price,o.name,o.description,s.section_key,
                o.selectable,o.display_order,o.active,o.display_behavior FROM option o
                JOIN section s ON s.id=o.section_id WHERE o.model_id=? ORDER BY o.sequence""",(mid,))]
            expected=[]
            for r in self.rows(roles["source_option_sheet"]):
                expected.append((r["option_id"],r["rpo"],None if r["price"] is None else str(r["price"]),r["option_name"],r["description"],r["section_id"],int(r["selectable"]),r["display_order"],int(r["active"]),r["display_behavior"]))
            self.assertEqual(actual,expected,key)
            statuses=[tuple(r) for r in self.db.execute("""SELECT o.legacy_id,v.legacy_id,a.status FROM availability a
                JOIN option o ON o.id=a.option_id JOIN variant v ON v.id=a.variant_id
                WHERE a.model_id=? ORDER BY a.sequence""",(mid,))]
            self.assertEqual(statuses,[(r["option_id"],r["variant_id"],r["status"]) for r in self.rows(roles["status_sheet"])],key)
        self.assertEqual(self.db.execute("SELECT count(*) FROM availability").fetchone()[0],7448)
        # Same RPO in the same model is deliberately two distinct options.
        rows=self.db.execute("""SELECT o.id FROM option o JOIN model m ON m.id=o.model_id
            WHERE m.model_key='grand_sport' AND o.rpo='T0E'""").fetchall()
        self.assertEqual(len({r[0] for r in rows}),2)

    def test_option_owner_preserves_price_semantics_and_source_links(self):
        self.assertEqual(self.db.execute("SELECT count(*) FROM option").fetchone()[0],1379)
        self.assertEqual(self.db.execute("SELECT count(*) FROM option WHERE rpo IS NULL").fetchone()[0],155)
        for option in self.db.execute("SELECT * FROM option"):
            self.assertEqual(option['price_basis'],'option')
            self.assertIsNone(option['currency'])
            self.assertEqual(option['rpo_role'], None if option['rpo'] is None else 'legacy-unspecified')
            source=self.db.execute("""SELECT s.ordinal,s.sheet FROM evidence_link e
                JOIN source_row s ON s.id=e.source_id WHERE e.entity_id=?""",(option['id'],)).fetchall()
            self.assertEqual(len(source),1)
            self.assertEqual(source[0]['ordinal'],option['sequence'])
            mapping=self.db.execute("SELECT kind,legacy_id,model_id FROM legacy_mapping WHERE entity_id=?",(option['id'],)).fetchone()
            self.assertEqual(tuple(mapping),('option',option['legacy_id'],option['model_id']))
        self.db.execute("UPDATE option SET base_price=NULL WHERE legacy_id='opt_r6x_001'")
        self.assertEqual(self.db.execute("SELECT count(*) FROM option WHERE legacy_id='opt_r6x_001' AND base_price IS NULL").fetchone()[0],6)
        self.db.execute("UPDATE option SET base_price='0' WHERE legacy_id='opt_r6x_001'")
        self.assertEqual(self.db.execute("SELECT count(*) FROM option WHERE legacy_id='opt_r6x_001' AND base_price='0'").fetchone()[0],6)

    def test_option_relationships_reject_cross_model_parents(self):
        row=self.db.execute("SELECT * FROM availability LIMIT 1").fetchone()
        other=self.db.execute("SELECT id FROM option WHERE model_id!=? LIMIT 1",(row['model_id'],)).fetchone()[0]
        self.db.execute("UPDATE availability SET option_id=? WHERE id=?",(other,row['id']))
        self.assertTrue(self.db.execute("PRAGMA foreign_key_check").fetchall())
        with self.assertRaises(sqlite3.IntegrityError):
            self.db.commit()

    def test_all_rule_endpoints_order_and_exact_amounts(self):
        for mid,key in self.db.execute("SELECT id,model_key FROM model"):
            roles={r["source_role"]:r["sheet_name"] for r in self.rows("model_workbook_sources") if r["model_key"]==key}
            for table,role,fields in [
                ("direct_rule","rule_mapping_sheet",("rule_id","source_id","target_id","rule_type","runtime_action","disabled_reason")),
                ("price_rule","price_rules_sheet",("price_rule_id","condition_option_id","target_option_id","price_rule_type","price_value"))]:
                left="source_id" if table=="direct_rule" else "condition_id"
                extra="r.runtime_action,r.explanation" if table=="direct_rule" else "r.amount"
                actual=[tuple(r) for r in self.db.execute(f"""SELECT r.legacy_id,l.legacy_id,t.legacy_id,r.effect,{extra}
                    FROM {table} r JOIN legacy_mapping l ON l.entity_id=r.{left}
                    JOIN legacy_mapping t ON t.entity_id=r.target_id WHERE r.model_id=? ORDER BY r.sequence""",(mid,))]
                expected=[]
                for r in self.rows(roles[role]):
                    v=[r[f] for f in fields]
                    if table=="price_rule":v[-1]=str(v[-1])
                    expected.append(tuple(v))
                self.assertEqual(actual,expected,(key,table))
        self.assertGreater(self.db.execute("SELECT count(*) FROM price_rule p JOIN model_interior i ON i.id=p.condition_id").fetchone()[0],0)

    def test_interior_paths_and_component_rate_fallback(self):
        for r in self.rows("model_interior_scope"):
            path=[x[0] for x in self.db.execute("""SELECT n.label FROM model_interior i JOIN model m ON m.id=i.model_id
                JOIN interior_hierarchy_member h ON h.interior_id=i.id JOIN hierarchy_node n ON n.id=h.node_id
                WHERE m.model_key=? AND i.legacy_id=? ORDER BY h.position""",(r["model_key"],r["interior_id"]))]
            self.assertEqual(path,json.loads(r["interior_hierarchy_levels"]))
        self.assertEqual(self.db.execute("SELECT count(*) FROM interior_component WHERE rate_id IS NULL").fetchone()[0],0)
        self.assertGreater(self.db.execute("""SELECT count(*) FROM interior_component c JOIN component_rate r ON r.id=c.rate_id
            WHERE c.rate_trim IS NOT NULL AND r.trim_scope=''""").fetchone()[0],0)

    def test_bucket_steps_and_shared_asset_scope_preserved(self):
        # Source labels newly exposed by D retain their actual routing-row provenance.
        for r in self.rows("model_workbook_sources"):
            if r["source_role"] not in {"source_option_sheet", "rule_mapping_sheet", "price_rules_sheet", "color_overrides_sheet", "interior_source_sheet"}:
                continue
            link=self.db.execute("""SELECT count(*) FROM evidence_link e JOIN source_row s ON s.id=e.source_id
                JOIN model_presentation p ON p.id=e.entity_id JOIN model m ON m.id=p.model_id
                WHERE m.model_key=? AND s.sheet='model_workbook_sources'""",(r["model_key"],)).fetchone()[0]
            self.assertEqual(link,5)
        self.assertEqual(self.db.execute("SELECT count(*) FROM runtime_step WHERE navigable=0 AND step_key='standard_equipment' AND runtime_order IS NULL").fetchone()[0],3)
        offered={(r[0],r[1]) for r in self.db.execute("SELECT model_id,legacy_id FROM option")}
        expected={(mid,r["target_id"],r["image_url"]) for r in self.rows("asset_map") if r["model_key"]=="*"
                  for (mid,) in self.db.execute("SELECT id FROM model") if (mid,r["target_id"]) in offered}
        actual={tuple(r) for r in self.db.execute("""SELECT a.model_id,o.legacy_id,a.image_url
            FROM asset_assignment a JOIN option o ON o.id=a.target_id WHERE a.source_scope='shared'""")}
        self.assertEqual(actual,expected)
        self.assertEqual(self.db.execute("SELECT count(*) FROM derivation_permission").fetchone()[0],5)

    def test_database_rejects_wrong_types_missing_and_cross_model_references(self):
        rule=self.db.execute("SELECT id,model_id FROM direct_rule LIMIT 1").fetchone()
        context=self.db.execute("SELECT id FROM context_choice WHERE model_id=? LIMIT 1",(rule["model_id"],)).fetchone()[0]
        with self.assertRaises(sqlite3.IntegrityError):
            self.db.execute("UPDATE direct_rule SET target_id=? WHERE id=?",(context,rule["id"]))
        with self.assertRaises(sqlite3.IntegrityError):
            self.db.execute("UPDATE direct_rule SET target_id=NULL WHERE id=?",(rule["id"],))
        with self.assertRaises(sqlite3.IntegrityError):
            self.db.execute("UPDATE option SET base_price='not money'")
        with self.assertRaises(sqlite3.IntegrityError):
            self.db.execute("UPDATE option SET selectable=2")
        other=self.db.execute("SELECT id FROM option WHERE model_id!=? LIMIT 1",(rule["model_id"],)).fetchone()[0]
        # Deferred ownership FK is checked at the transaction boundary.
        self.db.execute("UPDATE direct_rule SET target_id=? WHERE id=?",(other,rule["id"]))
        self.assertTrue(self.db.execute("PRAGMA foreign_key_check").fetchall())
        with self.assertRaises(sqlite3.IntegrityError):
            self.db.commit()

    def test_reimport_is_identical_and_modified_destination_is_preserved(self):
        before=self.path.read_bytes()
        self.assertEqual(build(self.path),self.report)
        self.assertEqual(self.path.read_bytes(),before)
        other=Path(self.temp.name)/"second.sqlite"
        build(other)
        self.assertEqual(other.read_bytes(),before)
        other.write_bytes(b"keep this existing file")
        with self.assertRaisesRegex(ValueError,"refusing to overwrite"):
            build(other)
        self.assertEqual(other.read_bytes(),b"keep this existing file")

    def test_failure_does_not_publish_or_leave_partial_database(self):
        target=Path(self.temp.name)/"failed.sqlite"
        with patch.object(Importer,"reconcile",side_effect=ValueError("forced reconciliation failure")):
            with self.assertRaisesRegex(ValueError,"forced reconciliation"):
                build(target)
        self.assertFalse(target.exists())
        self.assertFalse(list(target.parent.glob(".candidate-*")))
        with patch("catalog.importer.load_workbook",side_effect=ValueError("open failure")):
            with self.assertRaisesRegex(ValueError,"open failure"):
                build(target)
        self.assertFalse(target.exists())
        self.assertFalse(list(target.parent.glob(".candidate-*")))

    def test_missing_availability_is_rejected_even_with_valid_foreign_keys(self):
        wb=load_workbook(io.BytesIO(self.raw))
        wb["stingray_ovs"].delete_rows(2)
        db=connect(":memory:")
        create(db)
        try:
            with self.assertRaisesRegex(ValueError,"Missing availability pairs"):
                with db:Importer(db,wb).run()
            self.assertEqual(db.execute("SELECT count(*) FROM entity").fetchone()[0],0)
        finally:
            db.close();wb.close()

    def test_duplicate_legacy_identity_and_unmapped_fields_fail_closed(self):
        for kind in ("duplicate","new_field"):
            wb=load_workbook(io.BytesIO(self.raw))
            ws=wb["stingray_options"]
            if kind=="duplicate":ws.cell(3,1).value=ws.cell(2,1).value
            else:ws.cell(1,ws.max_column+1).value="unreviewed_field"
            db=connect(":memory:")
            create(db)
            try:
                with self.assertRaisesRegex(ValueError,"Duplicate|Unmapped fields"):
                    with db:Importer(db,wb).run()
            finally:
                db.close();wb.close()


if __name__=="__main__":
    unittest.main()
