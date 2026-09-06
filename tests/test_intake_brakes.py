"""Focused pilot checks; requires the separately supplied, unchanged raw guide."""
import io
import json
from pathlib import Path
import tarfile
import tempfile
import unittest

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

from scripts import intake_brakes as intake


class BrakeIntakeTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        # Missing local evidence must fail, never silently skip verification.
        cls.raw = intake.SOURCE.read_bytes()
        cls.archive = (intake.BASELINE / "workbook-runtime.tar.gz").read_bytes()
        with tarfile.open(fileobj=io.BytesIO(cls.archive)) as archive:
            cls.workbook_bytes = archive.extractfile("stingray_master.xlsx").read()
            cls.runtime_bytes = archive.extractfile(
                "form-output/runtime/stingray-runtime-contract.json").read()
        cls.guide = load_workbook(io.BytesIO(cls.raw), rich_text=True)
        cls.workbook = load_workbook(io.BytesIO(cls.workbook_bytes))
        cls.runtime = json.loads(cls.runtime_bytes)
        cls.result = intake.build()

    @classmethod
    def tearDownClass(cls):
        cls.guide.close()
        cls.workbook.close()

    def extract(self, guide=None, workbook=None):
        return intake.extract(guide or self.guide, workbook or self.workbook,
                              self.runtime, intake.digest(self.workbook_bytes),
                              intake.digest(self.runtime_bytes))

    def fresh_guide(self):
        # Reload original bytes: deepcopy loses openpyxl's custom format table.
        guide = load_workbook(io.BytesIO(self.raw), rich_text=True)
        self.addCleanup(guide.close)
        return guide

    def fresh_workbook(self):
        workbook = load_workbook(io.BytesIO(self.workbook_bytes))
        self.addCleanup(workbook.close)
        return workbook

    def test_all_eighteen_availability_pairs_match_inspected_evidence(self):
        expected_variants = ["1lt_c07", "2lt_c07", "3lt_c07", "1lt_c67", "2lt_c67", "3lt_c67"]
        pairs = [c for c in self.result["candidates"] if c["kind"] == "availability"]
        expected = {(variant, option): status for variant in expected_variants
                    for option, status in [("opt_jl9_001", "standard"),
                                           ("opt_j55_001", "available"),
                                           ("opt_z51_001", "available")]}
        actual = {(c["applicability"]["variant_id"], c["applicability"]["legacy_option_id"]):
                  c["typed_value"] for c in pairs}
        self.assertEqual(len(pairs), 18)
        self.assertEqual(actual, expected)
        self.assertTrue(all(c["comparison_class"] == "unchanged" for c in pairs))
        self.assertEqual([m["target_variant_id"] for m in self.result["variant_mappings"]], expected_variants)
        self.assertTrue(all(any(r.startswith("model_variants!") for r in m["baseline_evidence"])
                            for m in self.result["variant_mappings"]))

    def test_rich_footnotes_coverage_and_citations(self):
        evidence = self.result["evidence"]
        for column in "DEFGHI":
            runs = evidence[f"Mechanical 1!{column}7"]["runs"]
            self.assertEqual([(r["text"], r["vertical_align"]) for r in runs],
                             [("A", None), ("1", "superscript")])
            self.assertTrue(all(r["font_xml"] for r in runs))
        expected_cells = {f"{sheet}!{cell.coordinate}" for sheet, ranges in
                          [("Mechanical 1", intake.CONFIG["mechanical_ranges"]),
                           ("Price Schedule", intake.CONFIG["price_ranges"])]
                          for region in ranges for row in self.guide[sheet][region] for cell in row}
        self.assertEqual(set(evidence), expected_cells)
        for candidate in self.result["candidates"]:
            self.assertTrue(set(candidate["source_evidence"]) <= expected_cells)
        self.assertEqual(len(self.result["inventory"]), 28)
        self.assertEqual(sum(s["status"] == "out_of_scope" for s in self.result["inventory"]), 26)
        self.assertEqual(self.result["unresolved"]["external_package_codes"],
                         ["FE3", "G0K", "G96", "M1N", "QTU", "T0A", "V08"])
        self.assertFalse(self.result["unresolved"]["removal_authority"])
        self.assertFalse(self.result["unresolved"]["whole_guide_coverage"])

    def test_comparison_preserves_representation_and_price_ambiguities(self):
        by_kind = {c["kind"]: c for c in self.result["candidates"]}
        self.assertEqual(self.result["comparison_counts"], {"unchanged": 19, "ambiguous": 7})
        self.assertEqual(by_kind["includes"]["comparison_class"], "unchanged")
        self.assertEqual(by_kind["requires"]["baseline_value"],
                         {"explicit_rule_count": 0, "selectable": False})
        for kind in ["requires", "replaces", "code_role", "option_price", "base_price_context"]:
            self.assertEqual(by_kind[kind]["comparison_class"], "ambiguous")
        self.assertEqual(by_kind["option_price"]["typed_value"],
                         {"header_labeled_msrp": "0", "apparent_shifted_amount": "5395", "currency": None})
        base = by_kind["base_price_context"]
        self.assertEqual(base["typed_value"]["amount"], "73495")
        self.assertEqual(base["applicability"],
                         {"model": "stingray", "year": 2027, "variant_id": "1lt_c07"})
        self.assertEqual(self.result["accepted_operations"], [])
        self.assertTrue(all(c["status"] == "pending_review" for c in self.result["candidates"]))
        self.assertEqual(len(self.result["runtime_evidence"]), 1)
        self.assertEqual(self.result["runtime_evidence"][0]["value"]["target_id"], "opt_j55_001")

    def test_unrecognized_and_flattened_footnotes_remain_unresolved(self):
        for value in ["A1", "?", None, "A", CellRichText(
                TextBlock(InlineFont(), "A"), TextBlock(InlineFont(vertAlign="superscript"), "2"))]:
            with self.subTest(value=str(value)):
                guide = self.fresh_guide()
                guide["Mechanical 1"]["D7"] = value
                candidate = next(c for c in self.extract(guide=guide)["candidates"]
                                 if c["source_evidence"][0] == "Mechanical 1!D7")
                self.assertEqual(candidate["comparison_class"], "ambiguous")
                self.assertTrue(candidate["parse_issues"])
                if value in ["A1", "?", None]:
                    self.assertIsNone(candidate["typed_value"])

    def test_missing_note_does_not_infer_package_rules(self):
        guide = self.fresh_guide()
        guide["Mechanical 1"]["C7"] = "Unresolved note"
        result = self.extract(guide=guide)
        for candidate in result["candidates"]:
            if candidate["kind"] in {"includes", "requires"}:
                self.assertEqual(candidate["comparison_class"], "ambiguous")
                self.assertIsNone(candidate["typed_value"])

    def test_missing_duplicate_and_inactive_memberships_fail_closed(self):
        for change in ["missing", "duplicate", "inactive", "dangling", "inactive_variant"]:
            with self.subTest(change=change):
                workbook = self.fresh_workbook()
                members = workbook["model_variants"]
                if change == "missing":
                    members.delete_rows(14)
                elif change == "duplicate":
                    members.append([c.value for c in members[14]])
                elif change == "inactive":
                    members["D14"] = False
                elif change == "dangling":
                    members["B14"] = "missing_variant"
                else:
                    workbook["variant_master"]["H2"] = False
                with self.assertRaises(ValueError):
                    self.extract(workbook=workbook)

    def test_duplicate_variant_and_availability_identities_fail_closed(self):
        for sheet, row in [("variant_master", 2), ("stingray_ovs", 596), ("stingray_options", 168)]:
            with self.subTest(sheet=sheet):
                workbook = self.fresh_workbook()
                workbook[sheet].append([c.value for c in workbook[sheet][row]])
                with self.assertRaisesRegex(ValueError, "Expected one"):
                    self.extract(workbook=workbook)

    def test_wrong_or_duplicate_guide_headings_fail_closed(self):
        for heading in ["Unknown variant", str(self.guide["Mechanical 1"]["E3"].value)]:
            with self.subTest(heading=heading):
                guide = self.fresh_guide()
                guide["Mechanical 1"]["D3"] = heading
                with self.assertRaises(ValueError):
                    self.extract(guide=guide)

    def test_repeated_build_and_save_are_identical_and_do_not_overwrite(self):
        again = intake.build()
        self.assertEqual(intake.encoded(self.result), intake.encoded(again))
        ids = [c["id"] for c in again["candidates"]]
        self.assertEqual(len(ids), len(set(ids)))
        with tempfile.TemporaryDirectory() as directory:
            path, status = intake.save(self.result, directory)
            self.assertEqual(status, "created")
            stat = path.stat()
            self.assertEqual(intake.save(again, directory), (path, "reused"))
            self.assertEqual(path.stat().st_mtime_ns, stat.st_mtime_ns)
            self.assertEqual(list(Path(directory).iterdir()), [path])
            path.write_text("conflicting existing content")
            with self.assertRaisesRegex(ValueError, "refusing overwrite"):
                intake.save(again, directory)
            self.assertEqual(path.read_text(), "conflicting existing content")
        self.assertEqual(intake.SOURCE.read_bytes(), self.raw)
        self.assertEqual((intake.BASELINE / "workbook-runtime.tar.gz").read_bytes(), self.archive)

    def test_changed_baseline_creates_new_comparison_with_stable_candidates(self):
        changed = intake.extract(self.guide, self.workbook, self.runtime, "different-baseline",
                                 intake.digest(self.runtime_bytes))
        self.assertEqual(changed["run_id"], self.result["run_id"])
        self.assertNotEqual(changed["comparison_id"], self.result["comparison_id"])
        self.assertEqual([c["id"] for c in changed["candidates"]],
                         [c["id"] for c in self.result["candidates"]])

    def test_source_and_baseline_tampering_are_rejected(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "changed.xlsx"
            source.write_bytes(self.raw + b"changed")
            with self.assertRaisesRegex(ValueError, "Uninspected source revision"):
                intake.build(source)
            (root / "manifest.json").write_bytes((intake.BASELINE / "manifest.json").read_bytes())
            (root / "workbook-runtime.tar.gz").write_bytes(self.archive + b"changed")
            with self.assertRaisesRegex(ValueError, "Baseline archive hash mismatch"):
                intake.build(baseline=root)


if __name__ == "__main__":
    unittest.main()
